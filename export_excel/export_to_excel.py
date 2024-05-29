"""
Module for exporting data to Excel.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple
from pdf_reader.area import Area
from pdf_reader.pdf_parser import calculate_average


def export_to_excel(areas: Dict[str, Area], filename: str) -> None:
    """
    Export areas and their subjects to an Excel file.

    :param areas: Dictionary of Area objects.
    :param filename: The name of the Excel file to save the data.
    """
    data = prepare_data(areas)
    df = create_dataframe(data)
    save_dataframe_to_excel(df, filename)
    workbook, sheet = load_excel_workbook(filename)
    set_column_widths(sheet)
    align_headers(sheet)
    apply_borders(sheet)
    format_cells(sheet)
    apply_area_fill(sheet)
    weighted_grade, total_lp = calculate_average(areas)
    add_summary_table(sheet, total_lp, weighted_grade)
    workbook.save(filename)


def prepare_data(areas: Dict[str, Area]) -> List[List]:
    """
    Prepare data for the DataFrame.

    :param areas: Dictionary of Area objects.
    :return: List of lists containing the data.
    """
    data = []
    for area_name, area in areas.items():
        data.append([f"{area_name}", "", "", area.lp, area.calculate_total_lp(), round(area.calculate_total_weighted_sum() / area.calculate_total_lp(), 2)])
        for subject in area.subjects:
            data.append(["", subject.short, subject.name, subject.lp, subject.lp_evaluate, subject.grade])
    return data


def create_dataframe(data: List[List]) -> pd.DataFrame:
    """
    Create a DataFrame from the given data.

    :param data: List of lists containing the data.
    :return: DataFrame object.
    """
    return pd.DataFrame(data, columns=["Bereich", "Kürzel", "Name", "LP", "Angerechnete LP", "Note"])


def save_dataframe_to_excel(df: pd.DataFrame, filename: str) -> None:
    """
    Save the DataFrame to an Excel file.

    :param df: DataFrame object.
    :param filename: The name of the Excel file to save the data.
    """
    df.to_excel(filename, index=False, engine='openpyxl')


def load_excel_workbook(filename: str) -> Tuple[load_workbook, Worksheet]:
    """
    Load the Excel workbook and select the active worksheet.

    :param filename: The name of the Excel file.
    :return: Tuple containing the workbook and the active worksheet.
    """
    workbook = load_workbook(filename)
    sheet = workbook.active
    return workbook, sheet


def set_column_widths(sheet: Worksheet) -> None:
    """
    Set the width of each column in the worksheet.

    :param sheet: The active worksheet object.
    """
    column_widths = {
        "A": 30,  # Area
        "B": 15,  # Subject Code
        "C": 30,  # Subject Name
        "D": 6,   # LP
        "E": 13,  # Angerechnete LP
        "F": 6,   # Grade
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width


def align_headers(sheet: Worksheet) -> None:
    """
    Align headers to the upper left corner and wrap text.

    :param sheet: The active worksheet object.
    """
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)


def apply_borders(sheet: Worksheet) -> None:
    """
    Apply borders to all cells in the worksheet.

    :param sheet: The active worksheet object.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def format_cells(sheet: Worksheet) -> None:
    """
    Set the data types for specific columns and ensure text wrapping for the main part.

    :param sheet: The active worksheet object.
    """
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=6):  # Skip the header
        for cell in row:
            if cell.column_letter in ["D", "E"]:
                cell.number_format = "0"  # Integer format
            elif cell.column_letter == "F":
                cell.number_format = "0.0"  # Float format with 1 digit after the comma

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in ["A", "C", "E"]:
                cell.alignment = Alignment(wrap_text=True)


def apply_area_fill(sheet: Worksheet) -> None:
    """
    Apply the fill to the specific area_name rows.

    :param sheet: The active worksheet object.
    """
    blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=6, max_row=sheet.max_row), start=2):
        if not row[1].value:  # Check if the second column is empty, meaning it's an area_name row
            for cell in row:
                cell.fill = blue_fill


def add_summary_table(sheet: Worksheet, total_lp: int, weighted_sum: float) -> None:
    """
    Add a summary table with thin borders and green fill at the bottom of the worksheet.

    :param sheet: The active worksheet object.
    :param total_lp: Total LP value.
    :param weighted_sum: Total weighted sum value.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    data = [
        ("Total LP", total_lp),
        ("Total Sum", weighted_sum),
        ("Average", weighted_sum / total_lp)
    ]

    start_row = sheet.max_row + 2  # Start the summary table after the last row of data
    start_col = 1

    for i, (label, value) in enumerate(data):
        row = start_row + i
        sheet.cell(row=row, column=start_col, value=label).border = thin_border
        sheet.cell(row=row, column=start_col + 1, value=value).border = thin_border
        sheet.cell(row=row, column=start_col).fill = green_fill
        sheet.cell(row=row, column=start_col + 1).fill = green_fill